import logging
import re
from openpyxl import Workbook, load_workbook
from .entity.user import User
from .entity.tradeinfo import TradeInfo, SavingCardTradeInfo, CreditCardTradeInfo, CardType

logger = logging.getLogger(__name__)

# read
SHEET_SAVING_CARD = "large"
SHEET_CREDIT_CARD = "credit"

# create
SHEET_MERGE_CREATE = "output"

# title from large sheet 
TITLE_SAVING_CARD_USER_ID = "客户身份证件号码"
TITLE_SAVING_CARD_USER_NAME = "客户名称"
TITLE_SAVING_CARD_USER_NO = "客户号"
TITLE_SAVING_CARD_TRADE_DATE = "大额交易发生日期"
TITLE_SAVING_CARD_TRADE_AMOUNT = "交易金额"
TITLE_SAVING_CARD_TRADE_FINGER_PRINT = "大额交易特征代码"
TITLE_SAVING_CARD_TRADE_PURPOSE = "资金用途"
TITLE_SAVING_CARD_TRADE_TOTAL_NUM = "交易总数"

# title from credit sheet
TITLE_CREDIT_CARD_USER_ID = "客户身份证件"
TITLE_CREDIT_CARD_FLAG_IN_OR_OUT = "资金收付标识"
TITLE_CREDIT_CARD_FLAG_TRANSFER = "现金转账标识"
TITLE_CREDIT_CARD_TRADE_DATE = "交易日期"
TITLE_CREDIT_CARD_TRADE_AMOUNT = "交易金额"
TITLE_CREDIT_CARD_TRADE_PURPOSE = "资金用途"
TITLE_CREDIT_CARD_TRADE_STORE = "信用卡消费商户名称"
TITLE_CREDIT_CARD_TRADE_TYPE = "交易类型"

#客户身份证件号码 客户名称 客户号 大额交易发生日期 交易金额 大额交易特征代码 资金用途 交易总数 | 资金收付标识 现金转账标识 交易金额	资金用途 信用卡消费商户名称	交易类型
FIRST_ROW_WRITE = [TITLE_SAVING_CARD_USER_ID, TITLE_SAVING_CARD_USER_NAME, TITLE_SAVING_CARD_USER_NO, TITLE_SAVING_CARD_TRADE_DATE, TITLE_SAVING_CARD_TRADE_AMOUNT,
                   TITLE_SAVING_CARD_TRADE_FINGER_PRINT, TITLE_SAVING_CARD_TRADE_PURPOSE, TITLE_SAVING_CARD_TRADE_TOTAL_NUM,
                   TITLE_CREDIT_CARD_FLAG_IN_OR_OUT, TITLE_CREDIT_CARD_FLAG_TRANSFER, TITLE_CREDIT_CARD_TRADE_AMOUNT, TITLE_CREDIT_CARD_TRADE_PURPOSE,
                   TITLE_CREDIT_CARD_TRADE_STORE, TITLE_CREDIT_CARD_TRADE_TYPE]

RE_USER_ID = re.compile("[^0-9Xx]")

class BankExcelReader():
    def __init__(self, filePath, savePath):
        logger.debug("read file :%s save to: %s", filePath, savePath)
        self.file_path = filePath
        self.save_path = savePath
        self.work_book = load_workbook(filePath, read_only=True)
        # 身份证:User
        self.user_map = {}

    def getSavingCardSheet(self):
        return self.work_book[SHEET_SAVING_CARD]

    def getCreditCardSheet(self):
        return self.work_book[SHEET_CREDIT_CARD]

    def __getUser(self, identityId):
        return self.user_map[identityId]

    def __addUser(self, identityId, user):
        self.user_map[identityId] = user

    def __removeUser(self, id):
        self.user_map[id] = None

    def __readSheetNew(self, sheet, cardType):
        pass
    
    def __readSheet(self, sheet, cardType):
        first_row = sheet[1]

        for i in range(len(first_row)):
            if first_row[i].value != FIRST_ROW_READ[i]:
                raise Exception("列顺序不匹配")

        for row in sheet[2:sheet.max_row]:
            user_name = row[FIRST_ROW_READ.index(TITLE_USER_NAME)].value
            user_id = row[FIRST_ROW_READ.index(TITLE_USER_ID)].value
            user_id = RE_USER_ID.sub("", user_id)
            user = None

            if user_id in self.user_map:
                logger.debug("get user:%s", user_id)
                user = self.__getUser(user_id)
            else:
                if cardType == CardType.CARD_TYPE_CREDIT:
                    logger.debug("User[%s] Not in Saving Card Sheet Ignore!", user_id)
                else:
                    user = User(user_name, user_id)
                    self.__addUser(user_id, user)

            trade_time = row[FIRST_ROW_READ.index(TITLE_TRADE_DATE)].value
            trade_amount = row[FIRST_ROW_READ.index(TITLE_TRADE_AMOUNT)].value
            trade_type = row[FIRST_ROW_READ.index(TITLE_TRADE_TYPE)].value

            trade_info = TradeInfo(
                trade_time, trade_amount, trade_type, cardType)

            user.addTradeByTime(trade_time, trade_info)

    def __getRowIndexByValue(self, row, value):
        for i, cell in enumerate(row):
            if cell.value == value:
                return i
            
    def __readSavingCard(self):
        logger.debug("reading saving card sheet")

        saving_card_sheet = self.getSavingCardSheet()
        first_row = saving_card_sheet[1]
        
        for row in saving_card_sheet[2:saving_card_sheet.max_row]:
            user = None

            user_name = row[self.__getRowIndexByValue(first_row, TITLE_SAVING_CARD_USER_NAME)].value
            user_id = row[self.__getRowIndexByValue(first_row, TITLE_SAVING_CARD_USER_ID)].value
            user_id = RE_USER_ID.sub("", user_id)
            user_no = row[self.__getRowIndexByValue(first_row, TITLE_SAVING_CARD_USER_NO)].value
            
            if user_id in self.user_map:
                logger.debug("get user:%s", user_id)
                user = self.__getUser(user_id)
            else:
                user = User(user_name, user_id, user_no)
                self.__addUser(user_id, user)
                
            trade_time = row[self.__getRowIndexByValue(first_row, TITLE_SAVING_CARD_TRADE_DATE)].value
            trade_amount = row[self.__getRowIndexByValue(first_row, TITLE_SAVING_CARD_TRADE_AMOUNT)].value
            trade_finger_print = row[self.__getRowIndexByValue(first_row, TITLE_SAVING_CARD_TRADE_FINGER_PRINT)].value
            trade_purpose = row[self.__getRowIndexByValue(first_row, TITLE_SAVING_CARD_TRADE_PURPOSE)].value
            trade_total_num = row[self.__getRowIndexByValue(first_row, TITLE_SAVING_CARD_TRADE_TOTAL_NUM)].value
            
            trade_info = SavingCardTradeInfo(trade_time, trade_amount, trade_purpose,
                                              trade_finger_print,trade_total_num)
            
            user.addSavingCardTradeByTime(trade_time, trade_info)
            
    def __readCreditCard(self):
        logger.debug("reading credit card sheet")
        credit_card_sheet = self.getCreditCardSheet()
        first_row = credit_card_sheet[1]
        
        for row in credit_card_sheet[2:credit_card_sheet.max_row]:
            user = None

            user_id = row[self.__getRowIndexByValue(first_row, TITLE_CREDIT_CARD_USER_ID)].value
            user_id = RE_USER_ID.sub("", user_id)
            # 信用卡交易用户必须出现在大额交易中时才进行统计
            if user_id in self.user_map:
                logger.debug("get user:%s", user_id)
                user = self.__getUser(user_id)
            else:
                logger.warn("user %s not in saving card sheet ignore!!!", user_id)
                continue
                
            trade_time = row[self.__getRowIndexByValue(first_row, TITLE_CREDIT_CARD_TRADE_DATE)].value
            trade_amount = row[self.__getRowIndexByValue(first_row, TITLE_CREDIT_CARD_TRADE_AMOUNT)].value
            trade_purpose = row[self.__getRowIndexByValue(first_row, TITLE_CREDIT_CARD_TRADE_PURPOSE)].value
            trade_flag_in_out = row[self.__getRowIndexByValue(first_row, TITLE_CREDIT_CARD_FLAG_IN_OR_OUT)].value
            trade_flag_transfer = row[self.__getRowIndexByValue(first_row, TITLE_CREDIT_CARD_FLAG_TRANSFER)].value
            trade_store_name = row[self.__getRowIndexByValue(first_row, TITLE_CREDIT_CARD_TRADE_STORE)].value
            trade_type = row[self.__getRowIndexByValue(first_row, TITLE_CREDIT_CARD_TRADE_TYPE)].value

            trade_info = CreditCardTradeInfo(trade_time, trade_amount, trade_purpose,
                                            trade_flag_in_out,trade_flag_transfer,
                                            trade_store_name, trade_type)
            
            user.addCreditCardTradeInfo(trade_time, trade_info)

    def read(self):
        # must read saving card sheet first
        self.__readSavingCard()
        self.__readCreditCard()
        # close readonly workbook after read
        self.work_book.close()
        logger.debug("user map:", self.user_map)

    def __writeCell(self, sheet, user):
        for time, sc_trade_list in user.saving_card_trade_info.items():
            # 信用卡交易中包含同一时间的交易
            if time in user.credit_card_trade_info:
                # 先将大额交易写入表格
                for sc_trade_info in sc_trade_list:
                    sheet.append([user.identityId, user.name, user.userNo, time, sc_trade_info.amount, 
                                  sc_trade_info.finger_print, sc_trade_info.purpose, sc_trade_info.total_amount,
                                  None,None,None,None,None,None])

                # 再将同一时间的信用卡交易写入表格
                for cc_trade_info in user.credit_card_trade_info[time]:
                    sheet.append([user.identityId, user.name, user.userNo, time, None, None, None, None, 
                                  cc_trade_info.flag_in_or_out, cc_trade_info.flag_transfer, cc_trade_info.amount,
                                  cc_trade_info.purpose, cc_trade_info.store_name, cc_trade_info.trade_type])

            else:
                logger.debug("user:%s have no credit trade on %s", user.identityId, time)
                continue

    def write(self):
        save_work_book = Workbook(write_only=True)
        merge_sheet = save_work_book.create_sheet(SHEET_MERGE_CREATE)

        merge_sheet.append(FIRST_ROW_WRITE)

        for user in self.user_map.values():
            self.__writeCell(merge_sheet, user)
            self.__removeUser(user.identityId)

        save_work_book.save(self.save_path)
