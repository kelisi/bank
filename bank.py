import datetime
import logging
from openpyxl import Workbook, load_workbook

from excel.reader import BankExcelReader

logging.basicConfig(level=logging.DEBUG)

logger = logging.getLogger(__name__)


def w_write_tests():
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("sample.xlsx")


def w_read_test():
    logger.debug("ElSX READ TEST")
    wb = load_workbook("doc/dat.xlsx", read_only=True)

    for sheet in wb._sheets:
        logger.debug("Sheet :%s", sheet.title)

    ws = wb['大额交易']
    
    for row in ws.rows:
        for cell in row:
            print(cell.value)


def bank_read():
    logger.debug("============FBI WARING================")
    excel_reader = BankExcelReader("doc/dat.xlsx", "doc/out.xlsx")
    logger.debug("============Reading================")
    excel_reader.read()
    logger.debug("============Writing================")
    excel_reader.write()
    logger.debug("============EXIT================")


bank_read()
